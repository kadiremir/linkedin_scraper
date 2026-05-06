import argparse
from pathlib import Path

from linkedin_jobs.browser import collect_jobs_for_targets_from_linkedin
from linkedin_jobs.config import AppConfig, load_config, posted_within_seconds
from linkedin_jobs.models import JobPosting
from linkedin_jobs.profiles import resolve_browser_profile
from linkedin_jobs.urls import load_applied_job_ids


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = resolve_config(args)
    applied_job_ids = load_applied_job_ids(config.applied_jobs_file)
    if applied_job_ids:
        print(f"Excluding {len(applied_job_ids)} previously applied job(s).")
    browser_profile = resolve_browser_profile(args)
    headless = resolve_headless(args, config)
    if headless and args.pause_on_start:
        raise RuntimeError("--pause-on-start cannot be used with headless mode.")

    search_results = collect_jobs_for_targets_from_linkedin(
        user_data_dir=browser_profile.user_data_dir,
        chrome_profile_directory=browser_profile.profile_directory,
        pause_on_start=args.pause_on_start,
        targets=config.search_targets,
        max_pages=resolve_max_pages(args, config),
        filter_reposted=config.filter_reposted,
        posted_within_seconds=posted_within_seconds(config.posted_within),
        headless=headless,
        page_pause_seconds=args.page_pause_seconds,
        exclude_title_words=config.exclude_title_words,
        exclude_company_names=config.exclude_company_names,
        applied_job_ids=applied_job_ids,
    )

    titles_by_country: dict[str, list[str]] = {}
    for target in config.search_targets:
        titles_by_country.setdefault(target.country, []).append(target.job_title)

    total_jobs = sum(len(jobs) for _, jobs in search_results)
    if total_jobs == 0:
        print("No jobs found.")
        return 0

    print_all_job_details(search_results, titles_by_country)
    print(f"Total jobs found: {total_jobs}")
    write_jobs_to_excel(search_results, Path("output.xlsx"))
    return 0


def print_all_job_details(
    search_results: list[tuple[str, list[JobPosting]]],
    titles_by_country: dict[str, list[str]],
) -> None:
    for country, jobs in search_results:
        print_job_group(country, titles_by_country.get(country, []), jobs)


def print_job_group(country: str, job_titles: list[str], jobs: list[JobPosting]) -> None:
    roles = ", ".join(job_titles) if job_titles else "unknown"
    print(f"Searched for the roles: {roles} at: {country}")
    print(f"Found: {len(jobs)}")
    for index, job in enumerate(jobs, start=1):
        print(f"{index}. {format_job_summary(job)}")
        print(f"   link: {job.url}")
        print(f"   company: {job.company or ''}")
        print(f"   title: {job.title or ''}")
        print(f"   location: {job.location or ''}")
        print(f"   workplace_type: {job.workplace_type or ''}")
        print(f"   posted_at: {job.posted_at or ''}")


def write_jobs_to_excel(search_results: list[tuple[str, list[JobPosting]]], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = ["URL", "Country", "Company", "Title", "Location", "Workplace Type", "Posted At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    link_font = Font(color="0563C1", underline="single")

    for country, jobs in search_results:
        for job in jobs:
            ws.append([
                job.url,
                country,
                job.company or "",
                job.title or "",
                job.location or "",
                job.workplace_type or "",
                job.posted_at or "",
            ])
            url_cell = ws.cell(row=ws.max_row, column=1)
            url_cell.hyperlink = job.url
            url_cell.font = link_font

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(path)
    print(f"Results saved to {path}")


def format_job_summary(job: JobPosting) -> str:
    company = job.company or "Unknown company"
    title = job.title or "Unknown title"
    return f"{company} - {title}"


def resolve_config(args) -> AppConfig:
    if args.config.exists():
        return load_config(args.config)
    return AppConfig(countries=[args.location], job_titles=[args.keywords])


def resolve_max_pages(args, config: AppConfig) -> int | None:
    return args.max_pages if args.max_pages is not None else config.max_pages


def resolve_headless(args, config: AppConfig) -> bool:
    return args.headless or config.headless


def parse_args(argv: list[str] | None = None):
    parser = argparse.ArgumentParser(
        description="Find LinkedIn jobs and print parsed job details."
    )
    parser.add_argument("--keywords", default="software engineer")
    parser.add_argument("--location", default="Czechia")
    parser.add_argument("--config", type=Path, default=Path("config.yaml"))
    parser.add_argument("--profile-dir", type=Path, default=Path(".browser-profile"))
    parser.add_argument(
        "--use-default-chrome-profile",
        action="store_true",
        help="Use your Windows Chrome user data directory instead of the dedicated app profile.",
    )
    parser.add_argument(
        "--chrome-user-data-dir",
        type=Path,
        help="Chrome User Data directory to use, for example %%LOCALAPPDATA%%\\Google\\Chrome\\User Data.",
    )
    parser.add_argument(
        "--chrome-profile-directory",
        default="Default",
        help='Chrome profile folder inside the user data directory, for example "Default" or "Profile 1".',
    )
    parser.add_argument(
        "--pause-on-start",
        action="store_true",
        help="Pause after Chrome opens so you can manually confirm or fix the first tab.",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run Chrome without a visible window. Requires an already logged-in LinkedIn session.",
    )
    parser.add_argument(
        "--max-pages",
        type=positive_int,
        default=None,
        help="Optional page limit per country/title search. Defaults to config max_pages, or unlimited.",
    )
    parser.add_argument("--page-pause-seconds", type=non_negative_float, default=2.0)
    return parser.parse_args(argv)


def positive_int(value: str) -> int:
    parsed = int(value)
    if parsed < 1:
        raise argparse.ArgumentTypeError("must be at least 1")
    return parsed


def non_negative_float(value: str) -> float:
    parsed = float(value)
    if parsed < 0:
        raise argparse.ArgumentTypeError("must be greater than or equal to 0")
    return parsed
